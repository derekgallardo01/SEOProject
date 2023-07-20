from django.contrib import admin

from .models import ReportFiles

from django.forms import forms
from django.template import Template
from django.conf import settings
from django.core import serializers
import json
from django.core.files.storage import default_storage
from django.utils.safestring import mark_safe
from django.http import JsonResponse, HttpResponse,HttpResponseRedirect
import datetime
from django.utils import timezone
import time
import pandas as pd
import numpy as np
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color, Font, Fill, PatternFill, colors
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import YELLOW, RED
import random

class ReportFilesAdmin(admin.ModelAdmin):
    list_per_page = 20
    template_name = "profiles/user_profile.html"
    #fields = ['name','section']
    #formfield_overrides = ['name','section']
    fieldsets = (
        (None, {
            'fields': ('section','filename'),
        }),
    )
    list_display = ('name','section','all_actions')
    search_fields = ('name',)
    list_display_links = ('all_actions',)

    actions = ['merge_files', ]

    def all_actions(self,obj):
        return mark_safe('<span class="changelink">Edit</span>')
    all_actions.short_description = 'actions'

    class FilesInline(admin.StackedInline):
        model = ReportFiles
        extra = 1
        
    def save_model(self, request, obj, form, change):
        files = request.FILES.getlist('file_field')
        
        #return HttpResponse('ggg')
        for afile in request.FILES.getlist('files_multiple'):
            newdoc = ReportFiles(filename = afile, name = afile, section = request.POST["section"])
            newdoc.save()
            
            #obj.afile.path
            #obj.save()

    def merge_files(self, request, queryset):
        books = queryset.values_list('id','filename')
        
        timenow = timezone.now().strftime('%Y%m%d%H%M%S')
        # Python types will automatically be converted
        #excelfilename = settings.FILES_ROOT +'all_files_workbook'+ str(timenow)+'.xlsx'

        excelfilename = settings.FILES_ROOT +'all_files_workbook.xlsx'

        try:
            os.unlink(excelfilename)
        except:
            pass

        sno = 0
        sheetlist = []
        for file in books:
            sno +=1
            print(file[1])
            recid = file[0]
            csvdata = ReportFiles.objects.filter(pk=recid)
            
            filename = csvdata[0].filename
            #return HttpResponse('dddd')
            #sheetname = str(filename).replace('report_files/','')
            sheetname1 = csvdata[0].name
            rnn = random.randint(1,30)
            if(sheetname1 in sheetlist):
                sheetname = sheetname1+'_'+str(rnn)
            else:
                sheetname = sheetname1
            
            sheetlist.append(sheetname1)

            if(sheetname.find('crawl_overview') >=0):
                sheetname = sheetname.replace('.csv','')
                # load demo.xlsx 
                if os.path.isfile(excelfilename):
                    wb=load_workbook(excelfilename)
                else:
                    wb1 = Workbook()
                    wb1.save(excelfilename)
                    wb=load_workbook(excelfilename)
                
                wb.create_sheet(sheetname)
                ws = wb[sheetname]
                filename = settings.MEDIA_ROOT + str(filename)
                
                with open(filename) as f:
                    reader = csv.reader(f)
                    for row_index, row in enumerate(reader):
                        colored = 0
                        row_index += 1
                        msg = ''
                        #print(row_index)
                        if(len(row)>1 and row_index>6 and row[1].isnumeric() and int(row[1])>0 ):
                            #if(row[2] and (row[2])):
                            colored = 1
                            
                            

                        color = ''
                        msg = row[0].lower()
                        #print(msg) - 
                        #redlist = ['missing', 'duplicate', 'Client Error (4xx) response codes', '& URLs over 115 Characters', 'or Meta Description over 155 characters', 'contains canonical URLs', 'contains pagination', 'index and follow directives are present', 'all structured data is 100%']
                        redlist = ['Missing', 'Duplicate', 'Client Error (4xx)', 'Over 115 Characters', 'Over 155 Characters', 'Contains Canonical', 'Contains Pagination', 'Index', 'Follow','All']
                        if msg in [x.lower() for x in redlist]:
                            color = 'RED'
                        
                        #yellowlist = ['HTTP protocol currently active', 'redirection response codes', 'large pages over 100KB', 'non ASCII Characters URLs', 'Uppercase URLs', 'missing structured data', 'validation errors or warnings for structured data', 'Self-Referencing Canonicals', 'Canonicalized', 'Missing Canonicals']
                        yellowlist = ['Over 930 Pixels']
                        if msg in [x.lower() for x in yellowlist]:
                            color = 'YELLOW'

                        #orangelist = ['Meta Description Over 930 Pixels']
                        orangelist = ['HTTP', 'Redirection (3xx)', 'Large pages (over 100KB)', 'Non ASCII Characters', 'Uppercase', 'Validation Errors', 'Validation Warnings', 'Self Referencing', 'Canonicalised']
                        if msg in [x.lower() for x in orangelist]:
                            color = 'ORANGE'

                        #print(color)

                        for column_index, cell in enumerate(row):
                            column_letter = get_column_letter((column_index + 1))
                            #ws.cell('%s%s'%(column_letter, (row_index ))).value = cell
                            ws[('%s%s'%(column_letter, (row_index )))].value = cell
                            #print(cell)


                            if(colored==1):
                                if(color=='RED'):
                                    ws[('%s%s'%(column_letter, (row_index)))].fill = PatternFill(fgColor=RED, fill_type = "solid")
                                elif(color=='YELLOW'):
                                    ws[('%s%s'%(column_letter, (row_index)))].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
                                elif(color=='ORANGE'):
                                    ws[('%s%s'%(column_letter, (row_index)))].fill = PatternFill(start_color="FFA500", end_color="00FF00", fill_type = "solid")
                                else:
                                    pass
                                #ws[('%s%s'%(column_letter, (row_index)))].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                wb.save(excelfilename)
            else:
                pass

            # Save the file
            
        sheetlist = []
        for file in books:
            sno +=1
            print(file[1])
            recid = file[0]
            csvdata = ReportFiles.objects.filter(pk=recid)
            
            filename = csvdata[0].filename

            #return HttpResponse('dddd')
            #sheetname = str(filename).replace('report_files/','')

            sheetname1 = csvdata[0].name
            rnn = random.randint(1,30)
            if(sheetname1 in sheetlist):
                sheetname = sheetname1+str(rnn)
            else:
                sheetname = sheetname1
            

            if(sheetname.find('crawl_overview') <0):    
                sheetname = sheetname.replace('.csv','')
                
                # create new sheet
                print(sheetname)
                wb.create_sheet(sheetname)
                ws = wb[sheetname]
                filename = settings.MEDIA_ROOT + str(filename)
                
                with open(filename) as f:
                    reader = csv.reader(f)
                    for row in reader:
                        ws.append(row)

                # Save the file
                wb.save(excelfilename)

        #queryset.update(order_status=Order.CANCELLED)
        
        # load demo.xlsx 
        wb=load_workbook(excelfilename)
        # create new sheet
        wb.remove(wb.get_sheet_by_name('Sheet'))
        # save workbook
        wb.save(excelfilename)

        with open(excelfilename, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(excelfilename)
            return response

    merge_files.short_description = "Merged Files"


admin.site.register(ReportFiles,ReportFilesAdmin)

