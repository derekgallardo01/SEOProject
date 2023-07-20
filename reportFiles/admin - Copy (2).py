from django.contrib import admin

from .models import ReportFiles

from django.forms import forms
from django.template import Template
from django.conf import settings
from django.core import serializers
import json
from django.core.files.storage import default_storage
from django.utils.safestring import mark_safe
from django.http import JsonResponse, HttpResponse,HttpResponseRedirect
import datetime
from django.utils import timezone
import time
import pandas as pd
import numpy as np
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color, Font, Fill, PatternFill, colors


class ReportFilesAdmin(admin.ModelAdmin):
    list_per_page = 20
    template_name = "profiles/user_profile.html"
    #fields = ['name','section']
    #formfield_overrides = ['name','section']
    fieldsets = (
        (None, {
            'fields': ('section','filename'),
        }),
    )
    list_display = ('name','section','all_actions')
    search_fields = ('name',)
    list_display_links = ('all_actions',)

    actions = ['merge_files', ]

    def all_actions(self,obj):
        return mark_safe('<span class="changelink">Edit</span>')
    all_actions.short_description = 'actions'

    class FilesInline(admin.StackedInline):
        model = ReportFiles
        extra = 1
        
    def save_model(self, request, obj, form, change):
        files = request.FILES.getlist('file_field')
        
        #return HttpResponse('ggg')
        for afile in request.FILES.getlist('files_multiple'):
            newdoc = ReportFiles(filename = afile, name = afile, section = request.POST["section"])
            newdoc.save()
            
            #obj.afile.path
            #obj.save()

    def merge_files(self, request, queryset):
        books = queryset.values_list('id','filename')
        
        timenow = timezone.now().strftime('%Y%m%d%H%M%S')
        # Python types will automatically be converted
        excelfilename = settings.FILES_ROOT +'all_files_workbook'+ str(timenow)+'.xlsx'
        sno = 0
        for file in books:
            sno +=1
            print(file[1])
            recid = file[0]
            csvdata = ReportFiles.objects.filter(pk=recid)
            
            filename = csvdata[0].filename
            #return HttpResponse('dddd')
            sheetname = str(filename).replace('report_files/','')
            sheetname = sheetname.replace('.csv','')
            # load demo.xlsx 
            if os.path.isfile(excelfilename):
                wb=load_workbook(excelfilename)
            else:
                wb1 = Workbook()
                wb1.save(excelfilename)
                wb=load_workbook(excelfilename)
            # create new sheet
            print(sheetname)
            wb.create_sheet('Sheet-'+sheetname)
            #wb.sheetnames.append('Sheet-'+sheetname)
            #wb = Workbook()
            #ws = wb.sheetnames[sno]
            # grab the active worksheet
            ws = wb['Sheet-'+sheetname]
            filename = settings.MEDIA_ROOT + str(filename)



            RED = 'FFFF0000'
            with open(filename) as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
                    cd = ws.column_dimensions["A"]
                    for cell in ws["3:3"]:
                        #cell.font = red_font
                        #cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                        #cell.style.fill.fill_type = Fill.FILL_SOLID
                        cell.style.fill.start_color.index = colors.DARKRED
                    #cd.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                    #ws.cell(row=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                    #, column=2
                    #ws['A1'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type = "solid")
                
            # Save the file
            wb.save(excelfilename)

        #queryset.update(order_status=Order.CANCELLED)
        
        # load demo.xlsx 
        wb=load_workbook(excelfilename)
        # create new sheet
        wb.remove(wb.get_sheet_by_name('Sheet'))
        # save workbook
        wb.save(excelfilename)

        with open(excelfilename, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(excelfilename)
            return response

    merge_files.short_description = "Merged Files"


admin.site.register(ReportFiles,ReportFilesAdmin)

